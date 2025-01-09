import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def validate_belgian_vat(vat_number):
    """
    Validates a Belgian VAT number using a regular expression.
    Format: BE0xxx.xxx.xxx or 0xxx.xxx.xxx (with optional spaces or dots).
    """
    vat_pattern = r"^BE0\d{9}$|^0\d{9}$"
    return bool(re.match(vat_pattern, vat_number.replace(".", "").replace(" ", "")))


def highlight_errors_in_excel(file_path, sheet_name='Sheet1', vat_column='VAT Number',
                              output_file='highlighted_errors.xlsx'):
    """
    Reads an Excel file, validates VAT numbers, and highlights errors in the output file.

    Args:
        file_path (str): Path to the input Excel file.
        sheet_name (str): Name of the sheet to process.
        vat_column (str): Name of the column containing VAT numbers.
        output_file (str): Path to save the output Excel file.
    """
    # Load the Excel file into a DataFrame
    df =
    with pd.read_excel(file_path, sheet_name=sheet_name) as df:


    # Validate VAT numbers and create an 'Error' column
    df['Error'] = df[vat_column].apply(lambda vat: not validate_belgian_vat(str(vat)) if pd.notnull(vat) else True)

    # Save the DataFrame back to Excel
    df.to_excel(output_file, sheet_name=sheet_name, index=False, engine='openpyxl')

    # Load the workbook for highlighting
    wb = load_workbook(output_file)
    ws = wb[sheet_name]

    # Define a fill style for errors
    error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # Highlight errors in the VAT column
    for row_idx, is_error in enumerate(df['Error'], start=2):  # Start from the second row (headers are row 1)
        if is_error:
            ws[f"{vat_column[0]}{row_idx}"].fill = error_fill

    # Save the highlighted workbook
    wb.save(output_file)

# Example usage
# highlight_errors_in_excel('customers_suppliers.xlsx', vat_column='VAT Number')
